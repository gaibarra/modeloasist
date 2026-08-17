"""Service layer for superadmin attendance Excel uploads."""
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, time
from io import BytesIO
import re
from typing import Any
from unicodedata import normalize as unicode_normalize
from uuid import uuid4

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel as from_excel_datetime
from sqlalchemy import func, tuple_
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.dependencies.auth import AuthenticatedActor
from app.models.attendance_event import AttendanceEvent
from app.models.attendance_import_batch import AttendanceImportBatch
from app.models.employee import Employee
from app.models.employee_credential import EmployeeCredential
from app.models.staff_access import Department, DepartmentAlias, EmployeeDepartment
from app.schemas.staff import (
    AttendanceImportAutoCreatedEmployee,
    AttendanceImportBatchSummary,
    AttendanceImportDuplicateReason,
    AttendanceImportResult,
    AttendanceImportRowError,
)
from app.security import hash_password
from app.services.department_normalization import canonicalize_department, derive_department_campus


@dataclass
class ParsedAttendanceRow:
    row_number: int
    employee_id: int | None
    nombre: str
    departamento_raw: str
    fecha: date
    tiempo: time
    source: str | None
    device_name: str | None
    device_serial: str | None
    lookup_reason: str | None = None

    @property
    def event_ts(self) -> datetime:
        return datetime.combine(self.fecha, self.tiempo)

    @property
    def dedupe_key(self) -> tuple[int, date, time]:
        if self.employee_id is None:
            raise ValueError("employee_id no resuelto")
        return (self.employee_id, self.fecha, self.tiempo)


class AttendanceImportService:
    MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
    MAX_REPORTED_ERRORS = 50
    HEADER_SCAN_ROWS = 25

    HEADER_ALIASES = {
        "nombre": "nombre",
        "id": "employee_id",
        "employee id": "employee_id",
        "id de empleado": "employee_id",
        "codigo": "employee_id",
        "codigo de empleado": "employee_id",
        "codigo empleado": "employee_id",
        "matricula": "employee_id",
        "numero de empleado": "employee_id",
        "numero empleado": "employee_id",
        "num de empleado": "employee_id",
        "num empleado": "employee_id",
        "n": "employee_id",
        "n de empleado": "employee_id",
        "n o": "employee_id",
        "n o de empleado": "employee_id",
        "no de empleado": "employee_id",
        "no empleado": "employee_id",
        "enroll id": "employee_id",
        "enroll no": "employee_id",
        "enrol id": "employee_id",
        "employee no": "employee_id",
        "no": "employee_id",
        "no ": "employee_id",
        "numero": "employee_id",
        "num": "employee_id",
        "departamento": "departamento_raw",
        "fecha": "fecha",
        "date": "fecha",
        "tiempo": "tiempo",
        "hora": "tiempo",
        "time": "tiempo",
        "fuente de datos": "source",
        "fuente de datoss": "source",
        "nombre del dispositivo": "device_name",
        "no de serie del dispositivo": "device_serial",
        "n de serie del dispositivo": "device_serial",
        "n o de serie del dispositivo": "device_serial",
    }

    REQUIRED_FIELDS = {
        "employee_id",
        "nombre",
        "departamento_raw",
        "fecha",
        "tiempo",
        "source",
        "device_name",
        "device_serial",
    }

    DUPLICATE_REASON_LABELS = {
        "already_exists_in_database": "Ya existían en la base de datos",
        "duplicated_within_file": "Venían repetidos dentro del mismo archivo",
    }

    AUTO_CREATED_LOOKUP_LABELS = {
        "name_search_no_match": "Buscado por nombre sin éxito",
        "name_search_ambiguous": "Buscado por nombre con múltiples coincidencias",
        "employee_id_not_found": "ID no encontrado y sin coincidencia utilizable",
    }

    def __init__(self, *, db: Session, settings: Settings):
        self._db = db
        self._settings = settings

    def import_workbook(
        self,
        *,
        content: bytes,
        filename: str,
        actor: AuthenticatedActor,
    ) -> AttendanceImportResult:
        if actor.staff is None or not actor.is_superadmin:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acceso disponible solo para superadministradores")
        if not content:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo está vacío")
        if len(content) > self.MAX_FILE_SIZE_BYTES:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo excede el tamaño máximo permitido")

        rows = self._parse_workbook(content)
        if not rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo no contiene registros de asistencia")

        self._assign_missing_employee_ids(rows)
        auto_created_employees = self._ensure_employees(rows)
        imported_rows, skipped_duplicates, duplicate_breakdown, row_errors = self._persist_attendance_rows(rows)

        batch = AttendanceImportBatch(
            id=str(uuid4()),
            uploaded_by_staff_user_id=actor.staff.id,
            original_filename=filename,
            total_rows=len(rows),
            imported_rows=imported_rows,
            skipped_duplicates=skipped_duplicates,
            invalid_rows=len(row_errors),
            duplicate_breakdown=[item.model_dump() for item in duplicate_breakdown],
            auto_created_employees=[employee.model_dump() for employee in auto_created_employees],
        )
        self._db.add(batch)
        self._db.commit()
        self._db.refresh(batch)

        summary = AttendanceImportBatchSummary(
            id=batch.id,
            original_filename=batch.original_filename,
            uploaded_at=batch.uploaded_at,
            uploaded_by=actor.staff.full_name,
            total_rows=batch.total_rows,
            imported_rows=batch.imported_rows,
            skipped_duplicates=batch.skipped_duplicates,
            invalid_rows=batch.invalid_rows,
            duplicate_breakdown=duplicate_breakdown,
            auto_created_employees=auto_created_employees,
        )
        return AttendanceImportResult(batch=summary, row_errors=row_errors[: self.MAX_REPORTED_ERRORS])

    def _parse_workbook(self, content: bytes) -> list[ParsedAttendanceRow]:
        try:
            workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl exception types vary
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"No fue posible leer el archivo Excel: {exc}") from exc

        sheet = workbook.active
        sheet_rows = sheet.iter_rows(min_row=1, values_only=True)
        header_cells = next(sheet_rows, None)
        if header_cells is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo no contiene encabezados")

        preview_rows: list[tuple[int, list[Any]]] = []
        for row_number, values in enumerate(sheet_rows, start=2):
            preview_rows.append((row_number, list(values)))
            if len(preview_rows) >= self.HEADER_SCAN_ROWS:
                break

        header_map = self._resolve_header_map(list(header_cells), preview_rows)

        parsed_rows: list[ParsedAttendanceRow] = []
        row_errors: list[AttendanceImportRowError] = []
        for row_number, values in preview_rows:
            if values is None or all(value in (None, "") for value in values):
                continue
            try:
                parsed_rows.append(self._build_row(row_number=row_number, values=values, header_map=header_map))
            except ValueError as exc:
                row_errors.append(AttendanceImportRowError(row_number=row_number, message=str(exc)))

        for row_number, values in enumerate(sheet.iter_rows(min_row=len(preview_rows) + 2, values_only=True), start=len(preview_rows) + 2):
            if values is None or all(value in (None, "") for value in values):
                continue
            try:
                parsed_rows.append(self._build_row(row_number=row_number, values=list(values), header_map=header_map))
            except ValueError as exc:
                row_errors.append(AttendanceImportRowError(row_number=row_number, message=str(exc)))

        if not parsed_rows and row_errors:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "message": "No se pudieron procesar filas válidas del archivo",
                    "row_errors": [error.model_dump() for error in row_errors[: self.MAX_REPORTED_ERRORS]],
                },
            )

        self._parsing_errors = row_errors
        return parsed_rows

    def _resolve_header_map(self, headers: list[Any], preview_rows: list[tuple[int, list[Any]]]) -> dict[str, int]:
        mapping: dict[str, int] = {}
        employee_id_candidates: list[tuple[str, int]] = []
        name_candidates: list[tuple[str, int]] = []
        department_candidates: list[tuple[str, int]] = []
        date_candidates: list[tuple[str, int]] = []
        time_candidates: list[tuple[str, int]] = []
        max_columns = max([len(headers), *[len(values) for _, values in preview_rows]], default=len(headers))
        for index, header in enumerate(headers):
            normalized = self._normalize_header(header)
            field_name = self.HEADER_ALIASES.get(normalized)
            if field_name == "employee_id":
                employee_id_candidates.append((normalized, index))
                if field_name not in mapping:
                    mapping[field_name] = index
                continue
            if field_name == "nombre":
                name_candidates.append((normalized, index))
                if field_name not in mapping:
                    mapping[field_name] = index
                continue
            if field_name == "departamento_raw":
                department_candidates.append((normalized, index))
                if field_name not in mapping:
                    mapping[field_name] = index
                continue
            if field_name == "fecha":
                date_candidates.append((normalized, index))
                if field_name not in mapping:
                    mapping[field_name] = index
                continue
            if field_name == "tiempo":
                time_candidates.append((normalized, index))
                if field_name not in mapping:
                    mapping[field_name] = index
                continue
            if field_name and field_name not in mapping:
                mapping[field_name] = index
        mapping["employee_id"] = self._select_employee_id_column(employee_id_candidates, preview_rows, max_columns)
        mapping["fecha"] = self._select_date_column(date_candidates, preview_rows, max_columns)
        mapping["tiempo"] = self._select_time_column(time_candidates, preview_rows, max_columns, exclude_indices={mapping["fecha"]})
        excluded_indices = {
            mapping["employee_id"],
            mapping["fecha"],
            mapping["tiempo"],
            *[mapping[field_name] for field_name in ("source", "device_name", "device_serial") if field_name in mapping],
        }
        mapping["nombre"] = self._select_name_column(
            name_candidates,
            preview_rows,
            max_columns,
            exclude_indices=excluded_indices,
        )
        excluded_indices.add(mapping["nombre"])
        mapping["departamento_raw"] = self._select_department_column(
            department_candidates,
            preview_rows,
            max_columns,
            exclude_indices=excluded_indices,
        )
        missing_fields = sorted(self.REQUIRED_FIELDS - set(mapping))
        if missing_fields:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"El archivo no contiene las columnas requeridas: {', '.join(missing_fields)}",
            )
        return mapping

    def _select_date_column(
        self,
        candidates: list[tuple[str, int]],
        preview_rows: list[tuple[int, list[Any]]],
        max_columns: int,
    ) -> int:
        return self._select_column_by_content(
            candidates,
            preview_rows,
            self._looks_like_date,
            self._date_header_priority,
            max_columns=max_columns,
            missing_detail="No fue posible identificar las columnas de fecha y hora del archivo",
        )

    def _select_time_column(
        self,
        candidates: list[tuple[str, int]],
        preview_rows: list[tuple[int, list[Any]]],
        max_columns: int,
        exclude_indices: set[int] | None = None,
    ) -> int:
        return self._select_column_by_content(
            candidates,
            preview_rows,
            self._looks_like_time,
            self._time_header_priority,
            max_columns=max_columns,
            exclude_indices=exclude_indices,
            missing_detail="No fue posible identificar las columnas de fecha y hora del archivo",
        )

    def _select_name_column(
        self,
        candidates: list[tuple[str, int]],
        preview_rows: list[tuple[int, list[Any]]],
        max_columns: int,
        exclude_indices: set[int] | None = None,
    ) -> int:
        return self._select_column_by_content(
            candidates,
            preview_rows,
            self._looks_like_name,
            self._name_header_priority,
            max_columns=max_columns,
            exclude_indices=exclude_indices,
            missing_detail="No fue posible identificar la columna de nombre del archivo",
        )

    def _select_department_column(
        self,
        candidates: list[tuple[str, int]],
        preview_rows: list[tuple[int, list[Any]]],
        max_columns: int,
        exclude_indices: set[int] | None = None,
    ) -> int:
        return self._select_column_by_content(
            candidates,
            preview_rows,
            self._looks_like_department,
            self._department_header_priority,
            max_columns=max_columns,
            exclude_indices=exclude_indices,
            content_scorer=self._department_value_score,
            missing_detail="No fue posible identificar la columna de departamento del archivo",
        )

    def _select_column_by_content(
        self,
        candidates: list[tuple[str, int]],
        preview_rows: list[tuple[int, list[Any]]],
        validator: Any,
        priority_resolver: Any,
        *,
        max_columns: int,
        exclude_indices: set[int] | None = None,
        content_scorer: Any | None = None,
        missing_detail: str,
    ) -> int:
        excluded = exclude_indices or set()
        option_map: dict[int, str] = {index: normalized_header for normalized_header, index in candidates if index not in excluded}
        for index in range(max_columns):
            if index in excluded:
                continue
            option_map.setdefault(index, "")

        best_index: int | None = None
        best_score: tuple[int, int, int, int] | None = None
        for index, normalized_header in option_map.items():
            parseable = 0
            content_score = 0
            non_empty = 0
            for _, values in preview_rows:
                if index >= len(values):
                    continue
                value = values[index]
                if value in (None, ""):
                    continue
                non_empty += 1
                if validator(value):
                    parseable += 1
                    if content_scorer is not None:
                        content_score += content_scorer(value)
            score = (parseable, content_score, priority_resolver(normalized_header), non_empty)
            if best_score is None or score > best_score:
                best_score = score
                best_index = index
        if best_index is None or best_score is None or best_score[0] == 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=missing_detail)
        return best_index

    def _date_header_priority(self, normalized_header: str) -> int:
        if normalized_header in {"fecha", "date"}:
            return 2
        return 0

    def _time_header_priority(self, normalized_header: str) -> int:
        if normalized_header in {"tiempo", "hora", "time"}:
            return 2
        return 0

    def _name_header_priority(self, normalized_header: str) -> int:
        if normalized_header in {"nombre", "employee name", "name"}:
            return 2
        return 0

    def _department_header_priority(self, normalized_header: str) -> int:
        if normalized_header in {"departamento", "department"}:
            return 2
        return 0

    def _select_employee_id_column(
        self,
        candidates: list[tuple[str, int]],
        preview_rows: list[tuple[int, list[Any]]],
        max_columns: int,
    ) -> int:
        try:
            return self._select_column_by_content(
                candidates,
                preview_rows,
                self._looks_like_employee_id,
                self._employee_id_header_priority,
                max_columns=max_columns,
                missing_detail="No fue posible identificar la columna de numero de empleado del archivo",
            )
        except HTTPException:
            if not candidates:
                raise

        best_index = candidates[0][1]
        best_score: tuple[int, int] | None = None
        for normalized_header, index in candidates:
            non_empty = 0
            for _, values in preview_rows:
                if index >= len(values):
                    continue
                value = values[index]
                if value in (None, ""):
                    continue
                non_empty += 1
            score = (self._employee_id_header_priority(normalized_header), non_empty)
            if best_score is None or score > best_score:
                best_score = score
                best_index = index
        return best_index

    def _employee_id_header_priority(self, normalized_header: str) -> int:
        if normalized_header in {
            "employee id",
            "id de empleado",
            "codigo de empleado",
            "codigo empleado",
            "codigo",
            "matricula",
            "numero de empleado",
            "numero empleado",
            "num de empleado",
            "num empleado",
            "n",
            "n de empleado",
            "n o",
            "n o de empleado",
            "no de empleado",
            "no empleado",
            "enroll id",
            "enroll no",
            "enrol id",
            "employee no",
        }:
            return 3
        if normalized_header in {"no", "numero", "num"}:
            return 2
        if normalized_header == "id":
            return 1
        return 0

    def _looks_like_employee_id(self, value: Any) -> bool:
        try:
            self._coerce_int(value, "ID")
        except ValueError:
            return False
        return True

    def _looks_like_date(self, value: Any) -> bool:
        try:
            self._coerce_date(value, "Fecha")
        except ValueError:
            return False
        return True

    def _looks_like_time(self, value: Any) -> bool:
        try:
            self._coerce_time(value, "Tiempo")
        except ValueError:
            return False
        return True

    def _looks_like_name(self, value: Any) -> bool:
        text = self._coerce_optional_text(value)
        if not text:
            return False
        normalized_text = unicode_normalize("NFKC", text).strip()
        if "/" in normalized_text:
            return False
        if self._looks_like_date(normalized_text) or self._looks_like_time(normalized_text) or self._looks_like_employee_id(normalized_text):
            return False
        letters = sum(character.isalpha() for character in normalized_text)
        digits = sum(character.isdigit() for character in normalized_text)
        return letters >= 4 and letters > digits and " " in normalized_text

    def _looks_like_department(self, value: Any) -> bool:
        text = self._coerce_optional_text(value)
        if not text:
            return False
        normalized_text = unicode_normalize("NFKC", text).strip()
        if self._looks_like_date(normalized_text) or self._looks_like_time(normalized_text) or self._looks_like_employee_id(normalized_text):
            return False
        if re.fullmatch(r"[A-Z]{1,4}\d{4,}", normalized_text):
            return False
        letters = sum(character.isalpha() for character in normalized_text)
        digits = sum(character.isdigit() for character in normalized_text)
        return letters >= 4 and letters > digits and ("/" in normalized_text or "-" in normalized_text or " " in normalized_text)

    def _department_value_score(self, value: Any) -> int:
        text = self._coerce_optional_text(value)
        if not text:
            return 0
        normalized_text = unicode_normalize("NFKC", text).strip()
        score = 0
        if "/" in normalized_text:
            score += 3
        if normalized_text.count(" ") >= 1:
            score += 2
        if "-" in normalized_text:
            score += 1
        if len(normalized_text) >= 16:
            score += 1
        return score

    def _build_row(self, *, row_number: int, values: list[Any], header_map: dict[str, int]) -> ParsedAttendanceRow:
        employee_id = self._coerce_optional_int(self._get_value(values, header_map, "employee_id"), "ID")
        nombre = self._coerce_text(self._get_value(values, header_map, "nombre"), "nombre")
        departamento_raw = self._coerce_text(self._get_value(values, header_map, "departamento_raw"), "departamento")
        fecha_value = self._get_value(values, header_map, "fecha")
        tiempo_value = self._get_value(values, header_map, "tiempo")
        fecha, tiempo = self._coerce_date_time_pair(fecha_value=fecha_value, tiempo_value=tiempo_value)
        source = self._coerce_optional_text(self._get_value(values, header_map, "source"))
        device_name = self._coerce_optional_text(self._get_value(values, header_map, "device_name"))
        device_serial = self._coerce_optional_text(self._get_value(values, header_map, "device_serial"))
        return ParsedAttendanceRow(
            row_number=row_number,
            employee_id=employee_id,
            nombre=nombre,
            departamento_raw=departamento_raw,
            fecha=fecha,
            tiempo=tiempo,
            source=source,
            device_name=device_name,
            device_serial=device_serial,
        )

    def _coerce_date_time_pair(self, *, fecha_value: Any, tiempo_value: Any) -> tuple[date, time]:
        try:
            fecha = self._coerce_date(fecha_value, "Fecha")
            tiempo = self._coerce_time(tiempo_value, "Tiempo")
            return fecha, tiempo
        except ValueError as original_exc:
            try:
                swapped_fecha = self._coerce_date(tiempo_value, "Fecha")
                swapped_tiempo = self._coerce_time(fecha_value, "Tiempo")
                return swapped_fecha, swapped_tiempo
            except ValueError:
                raise original_exc

    def _assign_missing_employee_ids(self, rows: list[ParsedAttendanceRow]) -> None:
        existing_employees = self._db.query(Employee).all()
        existing_matches = {
            self._employee_identity_key_for_values(employee.nombre, employee.departamento): employee.id
            for employee in existing_employees
        }
        employees_by_name: dict[str, list[Employee]] = {}
        for employee in existing_employees:
            employees_by_name.setdefault(self._normalize_value(employee.nombre), []).append(employee)
        existing_ids = {employee.id for employee in existing_employees}

        for row in rows:
            matched_id = existing_matches.get(self._employee_identity_key(row))
            if matched_id is not None:
                row.employee_id = matched_id
                row.lookup_reason = None
                continue

            if row.employee_id is not None and row.employee_id in existing_ids:
                row.lookup_reason = None
                continue

            name_matches = employees_by_name.get(self._normalize_value(row.nombre), [])
            if len(name_matches) == 1:
                row.employee_id = name_matches[0].id
                row.lookup_reason = None
                continue

            if len(name_matches) > 1:
                row.lookup_reason = "name_search_ambiguous"
                continue

            row.lookup_reason = "name_search_no_match"

        missing_keys = {
            self._employee_identity_key(row): row
            for row in rows
            if row.employee_id is None
        }
        if not missing_keys:
            return

        reserved_ids = {row.employee_id for row in rows if row.employee_id is not None}
        next_employee_id = max(
            reserved_ids | {int(self._db.query(func.max(Employee.id)).scalar() or 0)},
            default=0,
        ) + 1
        generated_ids: dict[tuple[str, str], int] = {}

        for row in rows:
            if row.employee_id is not None:
                continue
            identity_key = self._employee_identity_key(row)
            matched_id = existing_matches.get(identity_key)
            if matched_id is not None:
                row.employee_id = matched_id
                continue
            assigned_id = generated_ids.get(identity_key)
            if assigned_id is None:
                while next_employee_id in reserved_ids:
                    next_employee_id += 1
                assigned_id = next_employee_id
                generated_ids[identity_key] = assigned_id
                reserved_ids.add(assigned_id)
                next_employee_id += 1
            row.employee_id = assigned_id

    def _ensure_employees(self, rows: list[ParsedAttendanceRow]) -> list[AttendanceImportAutoCreatedEmployee]:
        employee_ids = sorted({row.employee_id for row in rows if row.employee_id is not None})
        existing = {
            employee.id: employee
            for employee in self._db.query(Employee).filter(Employee.id.in_(employee_ids)).all()
        }
        auto_created: list[AttendanceImportAutoCreatedEmployee] = []
        rows_by_employee: dict[int, ParsedAttendanceRow] = {}
        for row in rows:
            if row.employee_id is None:
                continue
            rows_by_employee.setdefault(row.employee_id, row)

        for employee_id in employee_ids:
            row = rows_by_employee[employee_id]
            canonical_department = canonicalize_department(row.departamento_raw)
            existing_employee = existing.get(employee_id)
            if existing_employee is not None:
                if canonical_department.canonical_name and existing_employee.departamento != canonical_department.canonical_name:
                    existing_employee.departamento = canonical_department.canonical_name
                canonical_campus = canonical_department.campus or derive_department_campus(row.departamento_raw)
                if canonical_campus and existing_employee.campus != canonical_campus:
                    existing_employee.campus = canonical_campus
                self._ensure_department_link(employee=existing_employee, department_name=row.departamento_raw)
                continue
            email = f"emp-{employee_id}@pendiente.local"
            employee = Employee(
                id=employee_id,
                nombre=row.nombre,
                departamento=canonical_department.canonical_name,
                campus=canonical_department.campus or derive_department_campus(row.departamento_raw),
                division=self._extract_division(row.departamento_raw),
                email=email,
            )
            self._db.add(employee)
            self._db.flush()
            self._db.add(
                EmployeeCredential(
                    employee_id=employee.id,
                    password_hash=hash_password(self._settings.auth_default_password),
                    must_change_password=True,
                )
            )
            self._ensure_department_link(employee=employee, department_name=row.departamento_raw)
            auto_created.append(
                AttendanceImportAutoCreatedEmployee(
                    employee_id=employee.id,
                    nombre=employee.nombre,
                    departamento=employee.departamento,
                    email=employee.email,
                    lookup_reason=row.lookup_reason or "employee_id_not_found",
                    lookup_label=self.AUTO_CREATED_LOOKUP_LABELS.get(
                        row.lookup_reason or "employee_id_not_found",
                        self.AUTO_CREATED_LOOKUP_LABELS["employee_id_not_found"],
                    ),
                )
            )
        return auto_created

    def _persist_attendance_rows(
        self,
        rows: list[ParsedAttendanceRow],
    ) -> tuple[int, int, list[AttendanceImportDuplicateReason], list[AttendanceImportRowError]]:
        parsing_errors = list(getattr(self, "_parsing_errors", []))
        employee_ids = sorted({row.employee_id for row in rows if row.employee_id is not None})
        dates = sorted({row.fecha for row in rows})
        existing_keys = {
            (employee_id, fecha, tiempo)
            for employee_id, fecha, tiempo in self._db.query(
                AttendanceEvent.employee_id,
                AttendanceEvent.fecha,
                AttendanceEvent.tiempo,
            )
            .filter(AttendanceEvent.employee_id.in_(employee_ids), AttendanceEvent.fecha.in_(dates))
            .all()
        }
        imported_count = 0
        skipped_duplicates = 0
        duplicate_reason_counts = {
            "already_exists_in_database": 0,
            "duplicated_within_file": 0,
        }
        batch_seen: set[tuple[int, date, time]] = set()
        next_sqlite_event_id = self._next_bigint_id(AttendanceEvent) if self._db.bind and self._db.bind.dialect.name == "sqlite" else None

        for row in rows:
            if row.employee_id is None:
                parsing_errors.append(AttendanceImportRowError(row_number=row.row_number, message="No fue posible asignar un numero de empleado"))
                continue
            if row.dedupe_key in existing_keys:
                skipped_duplicates += 1
                duplicate_reason_counts["already_exists_in_database"] += 1
                continue
            if row.dedupe_key in batch_seen:
                skipped_duplicates += 1
                duplicate_reason_counts["duplicated_within_file"] += 1
                continue
            canonical_department = canonicalize_department(row.departamento_raw)
            event_kwargs: dict[str, Any] = {
                "employee_id": row.employee_id,
                "nombre": row.nombre,
                "departamento_raw": canonical_department.canonical_name or row.departamento_raw,
                "device_name": row.device_name,
                "device_serial": row.device_serial,
                "source": self._build_source_label(row.source),
                "fecha": row.fecha,
                "tiempo": row.tiempo,
                "event_ts": row.event_ts,
            }
            if next_sqlite_event_id is not None:
                event_kwargs["id"] = next_sqlite_event_id
                next_sqlite_event_id += 1
            self._db.add(AttendanceEvent(**event_kwargs))
            batch_seen.add(row.dedupe_key)
            imported_count += 1

        duplicate_breakdown = [
            AttendanceImportDuplicateReason(reason=reason, label=self.DUPLICATE_REASON_LABELS[reason], count=count)
            for reason, count in duplicate_reason_counts.items()
            if count > 0
        ]

        return imported_count, skipped_duplicates, duplicate_breakdown, parsing_errors

    def _ensure_department_link(self, *, employee: Employee, department_name: str) -> None:
        if not self._looks_like_department(department_name):
            return
        canonical_department = canonicalize_department(department_name)
        alias_values = [alias for alias in canonical_department.aliases if self._normalize_value(alias)]
        if not alias_values:
            return
        normalized_alias_values = [self._normalize_value(alias) for alias in alias_values]
        with self._db.no_autoflush:
            alias_rows = (
                self._db.query(DepartmentAlias.alias, DepartmentAlias.department_id)
                .filter(func.lower(DepartmentAlias.alias).in_(normalized_alias_values))
                .all()
            )
        alias_department_ids = {
            self._normalize_value(alias): int(department_id)
            for alias, department_id in alias_rows
            if self._normalize_value(alias)
        }
        for pending_alias in self._db.new:
            if not isinstance(pending_alias, DepartmentAlias):
                continue
            normalized_pending_alias = self._normalize_value(pending_alias.alias)
            if normalized_pending_alias in normalized_alias_values and pending_alias.department_id is not None:
                alias_department_ids[normalized_pending_alias] = int(pending_alias.department_id)
        canonical_alias = self._normalize_value(canonical_department.canonical_name)
        raw_alias = self._normalize_value(canonical_department.raw_name)
        department_id = alias_department_ids.get(canonical_alias) or alias_department_ids.get(raw_alias)
        department = self._db.get(Department, department_id) if department_id is not None else None
        if department is None:
            department = Department(
                id=self._next_bigint_id(Department) if self._is_sqlite() else None,
                code=self._build_department_code(canonical_department.canonical_name),
                name=canonical_department.canonical_name,
                campus=canonical_department.campus or derive_department_campus(department_name),
                active=True,
            )
            self._db.add(department)
            self._db.flush()
        else:
            canonical_name = canonical_department.canonical_name
            canonical_campus = canonical_department.campus or derive_department_campus(department_name)
            if canonical_name and department.name != canonical_name:
                department.name = canonical_name
            if canonical_campus and department.campus != canonical_campus:
                department.campus = canonical_campus
        with self._db.no_autoflush:
            existing_aliases = {
                self._normalize_value(alias)
                for (alias,) in self._db.query(DepartmentAlias.alias).filter(DepartmentAlias.department_id == department.id).all()
                if self._normalize_value(alias)
            }
        for pending_alias in self._db.new:
            if isinstance(pending_alias, DepartmentAlias) and pending_alias.department_id == department.id:
                normalized_pending_alias = self._normalize_value(pending_alias.alias)
                if normalized_pending_alias:
                    existing_aliases.add(normalized_pending_alias)
        next_alias_id = self._next_bigint_id(DepartmentAlias) if self._is_sqlite() else None
        for alias_value in canonical_department.aliases:
            normalized_alias = self._normalize_value(alias_value)
            if not normalized_alias or normalized_alias in existing_aliases:
                continue
            global_alias_owner_id = alias_department_ids.get(normalized_alias)
            if global_alias_owner_id is not None and global_alias_owner_id != department.id:
                continue
            alias = DepartmentAlias(
                id=next_alias_id,
                department_id=department.id,
                alias=alias_value,
                source="excel_import",
            )
            self._db.add(alias)
            existing_aliases.add(normalized_alias)
            alias_department_ids[normalized_alias] = int(department.id)
            if next_alias_id is not None:
                next_alias_id += 1
        has_link = (
            self._db.query(EmployeeDepartment)
            .filter(
                EmployeeDepartment.employee_id == employee.id,
                EmployeeDepartment.department_id == department.id,
            )
            .first()
        )
        if has_link is None:
            self._db.add(EmployeeDepartment(employee_id=employee.id, department_id=department.id, is_primary=True))

    def _get_value(self, values: list[Any], header_map: dict[str, int], field_name: str) -> Any:
        index = header_map[field_name]
        return values[index] if index < len(values) else None

    def _coerce_int(self, value: Any, label: str) -> int:
        if value in (None, ""):
            raise ValueError(f"{label} es obligatorio")
        if isinstance(value, (int, float)):
            integer_value = int(value)
            if integer_value <= 0:
                raise ValueError(f"{label} debe ser un entero positivo")
            return integer_value
        text = unicode_normalize("NFKC", str(value)).strip().replace("'", "").replace("\u2019", "")
        if not text:
            raise ValueError(f"{label} es obligatorio")
        while text.startswith("="):
            text = text[1:].strip()
        if len(text) >= 2 and text[0] == text[-1] and text[0] in {'"', "'"}:
            text = text[1:-1].strip()
        compact_text = re.sub(r"\s+", "", text)
        try:
            integer_value = self._parse_integer_text(compact_text, label)
        except ValueError as exc:
            if str(exc) == f"{label} debe ser numérico":
                raise ValueError(f"{label} debe ser numérico (valor: {self._preview_value(value)})") from exc
            raise
        if integer_value <= 0:
            raise ValueError(f"{label} debe ser un entero positivo")
        return integer_value

    def _coerce_optional_int(self, value: Any, label: str) -> int | None:
        if value in (None, ""):
            return None
        try:
            return self._coerce_int(value, label)
        except ValueError:
            return None

    def _parse_integer_text(self, text: str, label: str) -> int:
        if text.isdigit():
            return int(text)

        if re.fullmatch(r"\d{1,3}(?:,\d{3})+", text) or re.fullmatch(r"\d{1,3}(?:\.\d{3})+", text):
            return int(text.replace(",", "").replace(".", ""))

        if re.fullmatch(r"\d{1,3}(?:,\d{3})+\.0+", text):
            integer_portion = text.split(".", 1)[0]
            return int(integer_portion.replace(",", ""))

        if re.fullmatch(r"\d{1,3}(?:\.\d{3})+,0+", text):
            integer_portion = text.split(",", 1)[0]
            return int(integer_portion.replace(".", ""))

        if re.fullmatch(r"\d+[\.,]0+", text):
            integer_portion = text.split(",", 1)[0].split(".", 1)[0]
            return int(integer_portion)

        normalized_decimal = text
        if "," in normalized_decimal and "." not in normalized_decimal and not re.fullmatch(r"\d{1,3}(?:,\d{3})+", normalized_decimal):
            normalized_decimal = normalized_decimal.replace(",", ".")
        try:
            decimal_value = Decimal(normalized_decimal)
        except InvalidOperation as exc:
            raise ValueError(f"{label} debe ser numérico") from exc
        if decimal_value != decimal_value.to_integral_value():
            raise ValueError(f"{label} debe ser numérico")
        return int(decimal_value)

    def _coerce_text(self, value: Any, label: str) -> str:
        text = self._coerce_optional_text(value)
        if not text:
            raise ValueError(f"{label} es obligatorio")
        return text

    def _coerce_optional_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _coerce_date(self, value: Any, label: str) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            try:
                excel_value = from_excel_datetime(value)
            except Exception:
                excel_value = None
            if isinstance(excel_value, datetime):
                return excel_value.date()
            if isinstance(excel_value, date):
                return excel_value
        text = self._coerce_text(value, label)
        normalized_text = unicode_normalize("NFKC", text).strip()
        date_only_text = re.split(r"[T\s]+", normalized_text, maxsplit=1)[0]
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(normalized_text, fmt).date()
            except ValueError:
                try:
                    return datetime.strptime(date_only_text, fmt).date()
                except ValueError:
                    continue
        raise ValueError(f"{label} tiene un formato inválido (valor: {self._preview_value(value)})")

    def _coerce_time(self, value: Any, label: str) -> time:
        if isinstance(value, datetime):
            return value.time().replace(microsecond=0)
        if isinstance(value, time):
            return value.replace(microsecond=0)
        text = self._coerce_text(value, label)
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                continue
        raise ValueError(f"{label} tiene un formato inválido")

    def _normalize_header(self, value: Any) -> str:
        if value is None:
            return ""
        text = unicode_normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
        text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
        return re.sub(r"\s+", " ", text)

    def _normalize_value(self, value: str | None) -> str:
        return (value or "").strip().lower()

    def _employee_identity_key(self, row: ParsedAttendanceRow) -> tuple[str, str]:
        return self._employee_identity_key_for_values(row.nombre, row.departamento_raw)

    def _employee_identity_key_for_values(self, nombre: str | None, departamento_raw: str | None) -> tuple[str, str]:
        canonical_department = canonicalize_department(departamento_raw)
        normalized_department = self._normalize_value(canonical_department.canonical_name or canonical_department.raw_name)
        return (self._normalize_value(nombre), normalized_department)

    def _preview_value(self, value: Any) -> str:
        text = unicode_normalize("NFKC", str(value)).replace("\r", " ").replace("\n", " ").strip()
        compact_text = re.sub(r"\s+", " ", text)
        if len(compact_text) > 32:
            return f"{compact_text[:29]}..."
        return compact_text or "<vacío>"

    def _extract_campus(self, department_name: str | None) -> str | None:
        return derive_department_campus(department_name)

    def _extract_division(self, department_name: str | None) -> str | None:
        if not department_name:
            return None
        parts = [chunk.strip() for chunk in department_name.split("/") if chunk.strip()]
        if parts:
            return parts[0]
        return None

    def _build_department_code(self, department_name: str) -> str:
        base_code = re.sub(r"[^a-zA-Z0-9]+", "-", self._normalize_value(department_name)).strip("-") or "department"
        candidate = base_code[:255]
        counter = 1
        while self._db.query(Department).filter(Department.code == candidate).first() is not None:
            suffix = f"-{counter}"
            candidate = f"{base_code[: max(1, 255 - len(suffix))]}{suffix}"
            counter += 1
        return candidate

    def _build_source_label(self, source: str | None) -> str:
        base = (source or "excel").strip()
        label = f"excel:{base}"
        return label[:64]

    def _next_bigint_id(self, model: type[Any]) -> int:
        max_id = self._db.query(func.max(model.id)).scalar()
        return int(max_id or 0) + 1

    def _is_sqlite(self) -> bool:
        return bool(self._db.bind and self._db.bind.dialect.name == "sqlite")