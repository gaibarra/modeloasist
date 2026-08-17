from datetime import date, datetime, time
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app
from app.models.attendance_event import AttendanceEvent
from app.models.attendance_import_batch import AttendanceImportBatch
from app.models.employee import Employee
from app.models.employee_credential import EmployeeCredential
from app.models.staff_access import Department, DepartmentAlias, EmployeeDepartment, StaffDepartmentScope, StaffUser
from app.security import hash_password
from tests.conftest import TestingSessionLocal


def seed_attendance_import_data(*, is_superadmin: bool) -> None:
    session = TestingSessionLocal()
    session.query(AttendanceImportBatch).delete()
    session.query(AttendanceEvent).delete()
    session.query(StaffDepartmentScope).delete()
    session.query(StaffUser).delete()
    session.query(EmployeeDepartment).delete()
    session.query(DepartmentAlias).delete()
    session.query(Department).delete()
    session.query(EmployeeCredential).delete()
    session.query(Employee).delete()

    department_name = "Escuela Modelo/Montejo/MTJ-Imagen"
    employee = Employee(
        id=3074,
        nombre="CANCHE MAY ROGER",
        departamento=department_name,
        campus="Montejo",
        email="cancheroger@example.com",
    )
    department = Department(id=3, code="mtj-imagen", name=department_name, campus="Montejo", active=True)
    alias = DepartmentAlias(id=1, department_id=3, alias=department_name, source="employee")
    employee_department = EmployeeDepartment(employee_id=3074, department_id=3, is_primary=True)
    credential = EmployeeCredential(employee_id=3074, password_hash=hash_password("modelo2026"), must_change_password=False)
    existing_event = AttendanceEvent(
        id=1,
        employee_id=3074,
        nombre="CANCHE MAY ROGER",
        departamento_raw=department_name,
        device_name="CVA EDIFICIO PRINCIPAL",
        device_serial="FQ8321320",
        source="excel:Dispositivo",
        fecha=date(2026, 4, 19),
        tiempo=time(6, 1),
        event_ts=datetime(2026, 4, 19, 6, 1),
        created_at=datetime(2026, 4, 19, 6, 1),
    )
    staff = StaffUser(
        id=10,
        email="staff@example.com",
        full_name="Staff Demo",
        password_hash=hash_password("staff2026!"),
        must_change_password=False,
        is_active=True,
        is_superadmin=is_superadmin,
        employee_id=None,
    )
    scope = StaffDepartmentScope(staff_user_id=10, department_id=3)

    session.add_all([employee, department, alias, employee_department, credential, existing_event, staff, scope])
    session.commit()
    session.close()


def auth_headers(client: TestClient) -> dict[str, str]:
    response = client.post("/auth/login", json={"email": "staff@example.com", "password": "staff2026!"})
    assert response.status_code == 200
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def build_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "CANCHE MAY ROGER",
        3074,
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "19-04-2026",
        "06:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "REJON CAPILLA CARMEN DOLORES",
        5005789865,
        "CME-Soluciones Castro",
        "19-04-2026",
        "07:17",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_duplicate_rows_in_same_file() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    repeated_row = [
        "CANCHE MAY ROGER",
        3074,
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "20-04-2026",
        "06:01",
        "Lunes",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ]
    sheet.append(repeated_row)
    sheet.append(repeated_row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_text_ids() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "CANCHE MAY ROGER",
        "3,074",
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "19-04-2026",
        "06:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "REJON CAPILLA CARMEN DOLORES",
        "5.005789865E+9",
        "CME-Soluciones Castro",
        "19-04-2026",
        "07:17",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_localized_ids() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "CANCHE MAY ROGER",
        "3.074",
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "19-04-2026",
        "06:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "REJON CAPILLA CARMEN DOLORES",
        "5,005,789,865",
        "CME-Soluciones Castro",
        "19-04-2026",
        "07:17",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_grouped_decimal_ids() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "CANCHE MAY ROGER",
        "3,074.00",
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "19-04-2026",
        "06:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "REJON CAPILLA CARMEN DOLORES",
        "5.005.789.865,00",
        "CME-Soluciones Castro",
        "19-04-2026",
        "07:17",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_alternate_employee_number_column() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "No.",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "CANCHE MAY ROGER",
        3074,
        "MTJ-Imagen",
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "19-04-2026",
        "06:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "REJON CAPILLA CARMEN DOLORES",
        5005789865,
        "MTJ-Soluciones Castro",
        "CME-Soluciones Castro",
        "19-04-2026",
        "07:17",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_without_numeric_employee_ids() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "NUEVA PERSONA UNO",
        "MTJ-Imagen",
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "19-04-2026",
        "06:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "NUEVA PERSONA UNO",
        "MTJ-Imagen",
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "19-04-2026",
        "14:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "NUEVA PERSONA DOS",
        "MTJ-Soluciones Castro",
        "CME-Soluciones Castro",
        "19-04-2026",
        "07:17",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_datetime_date_values() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "CANCHE MAY ROGER",
        3074,
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "2026-04-19 06:01:00",
        "06:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "REJON CAPILLA CARMEN DOLORES",
        5005789865,
        "CME-Soluciones Castro",
        "19/04/2026 07:17",
        "07:17",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_swapped_date_time_columns() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "CANCHE MAY ROGER",
        3074,
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "06:01",
        "19-04-2026",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "REJON CAPILLA CARMEN DOLORES",
        5005789865,
        "CME-Soluciones Castro",
        "07:17",
        "19-04-2026",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_hidden_real_date_column() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Transacción",
        "Día laborable",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "CANCHE MAY ROGER",
        3074,
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "06:01",
        "Domingo",
        "19-04-2026",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    sheet.append([
        "REJON CAPILLA CARMEN DOLORES",
        5005789865,
        "CME-Soluciones Castro",
        "07:17",
        "Domingo",
        "19-04-2026",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_misleading_named_columns() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "ID",
        "nombre",
        "departamento",
        "Fecha",
        "Tiempo",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
        "No. empleado real",
        "Persona registrada",
        "Ubicación laboral",
    ])
    sheet.append([
        "MTJ-Imagen",
        "3074",
        "19-04-2026",
        "06:01",
        "Domingo",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
        3074,
        "CANCHE MAY ROGER",
        "Escuela Modelo/Montejo/MTJ-Imagen",
    ])
    sheet.append([
        "MTJ-Soluciones Castro",
        "5005789865",
        "19-04-2026",
        "07:17",
        "Domingo",
        "Dispositivo",
        "CME EXTERNOS",
        "FQ2661487",
        5005789865,
        "REJON CAPILLA CARMEN DOLORES",
        "CME-Soluciones Castro",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_for_existing_employee_name_match() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "AKE GARCIA JUAN CARLOS",
        3095,
        "MTJ-Imagen",
        "19-04-2026",
        "07:30",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_with_existing_employee_under_different_file_id() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "AKE GARCIA JUAN CARLOS",
        3095,
        "MTJ-Imagen",
        "19-04-2026",
        "08:30",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_for_existing_employee_name_match_with_long_department() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "AKE GARCIA JUAN CARLOS",
        "MTJ-Imagen",
        "Escuela Modelo/Montejo/MTJ-Imagen",
        "19-04-2026",
        "07:30",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_workbook_bytes_for_generic_prefixed_department() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "PERSONA BAJAS",
        8001,
        "Escuela Modelo/BAJAS",
        "19-04-2026",
        "07:30",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_superadmin_can_import_attendance_and_auto_create_employee():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["invalid_rows"] == 0
    assert payload["batch"]["duplicate_breakdown"] == [
        {
            "reason": "already_exists_in_database",
            "label": "Ya existían en la base de datos",
            "count": 1,
        }
    ]
    assert payload["batch"]["auto_created_employees"] == [
        {
            "employee_id": 5005789865,
            "nombre": "REJON CAPILLA CARMEN DOLORES",
            "departamento": "CME-Soluciones Castro",
            "email": "emp-5005789865@pendiente.local",
            "lookup_reason": "name_search_no_match",
            "lookup_label": "Buscado por nombre sin éxito",
        }
    ]

    session = TestingSessionLocal()
    created_employee = session.query(Employee).filter(Employee.id == 5005789865).one()
    created_credential = session.query(EmployeeCredential).filter(EmployeeCredential.employee_id == 5005789865).one()
    created_event = (
        session.query(AttendanceEvent)
        .filter(AttendanceEvent.employee_id == 5005789865, AttendanceEvent.fecha == date(2026, 4, 19))
        .one()
    )
    batch = session.query(AttendanceImportBatch).one()
    session.close()

    assert created_employee.email == "emp-5005789865@pendiente.local"
    assert created_credential.must_change_password is True
    assert created_event.device_serial == "FQ2661487"
    assert batch.imported_rows == 1


def test_non_superadmin_cannot_import_attendance():
    seed_attendance_import_data(is_superadmin=False)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 403


def test_list_attendance_import_history_returns_latest_batch():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)
    headers = auth_headers(client)
    upload_response = client.post(
        "/staff/attendance-imports",
        headers=headers,
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 201

    response = client.get("/staff/attendance-imports", headers=headers)

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 1
    assert payload[0]["original_filename"] == "asistencia.xlsx"
    assert payload[0]["uploaded_by"] == "Staff Demo"
    assert payload[0]["duplicate_breakdown"] == [
        {
            "reason": "already_exists_in_database",
            "label": "Ya existían en la base de datos",
            "count": 1,
        }
    ]
    assert payload[0]["auto_created_employees"][0]["employee_id"] == 5005789865
    assert payload[0]["auto_created_employees"][0]["lookup_reason"] == "name_search_no_match"


def test_superadmin_reports_duplicates_repeated_within_same_file():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_duplicate_rows_in_same_file(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["duplicate_breakdown"] == [
        {
            "reason": "duplicated_within_file",
            "label": "Venían repetidos dentro del mismo archivo",
            "count": 1,
        }
    ]
    assert payload["batch"]["auto_created_employees"] == []


def test_superadmin_can_import_attendance_with_excel_text_ids():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_text_ids(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["invalid_rows"] == 0


def test_superadmin_can_import_attendance_with_localized_grouped_ids():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_localized_ids(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["invalid_rows"] == 0


def test_superadmin_can_import_attendance_with_grouped_decimal_ids():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_grouped_decimal_ids(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    assert response.json()["batch"]["imported_rows"] == 1
    assert response.json()["batch"]["skipped_duplicates"] == 1


def test_superadmin_can_import_attendance_when_plain_id_column_is_department_code():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_alternate_employee_number_column(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["invalid_rows"] == 0


def test_superadmin_assigns_employee_numbers_when_ids_are_not_numeric():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_without_numeric_employee_ids(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 3
    assert payload["batch"]["skipped_duplicates"] == 0
    assert payload["batch"]["invalid_rows"] == 0
    created_ids = [item["employee_id"] for item in payload["batch"]["auto_created_employees"]]
    assert created_ids == [3075, 3076]

    session = TestingSessionLocal()
    generated_employee = session.query(Employee).filter(Employee.id == 3075).one()
    generated_events = (
        session.query(AttendanceEvent)
        .filter(AttendanceEvent.employee_id == 3075)
        .order_by(AttendanceEvent.tiempo.asc())
        .all()
    )
    second_employee = session.query(Employee).filter(Employee.id == 3076).one()
    session.close()

    assert generated_employee.nombre == "NUEVA PERSONA UNO"
    assert len(generated_events) == 2
    assert second_employee.nombre == "NUEVA PERSONA DOS"


def test_superadmin_can_import_attendance_with_datetime_date_strings():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_datetime_date_values(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["invalid_rows"] == 0


def test_superadmin_can_import_attendance_when_date_and_time_columns_are_swapped():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_swapped_date_time_columns(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["invalid_rows"] == 0


def test_superadmin_can_import_attendance_when_real_date_column_uses_unknown_header():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_hidden_real_date_column(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["invalid_rows"] == 0


def test_superadmin_can_import_attendance_when_name_and_department_headers_are_misleading():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_misleading_named_columns(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["skipped_duplicates"] == 1
    assert payload["batch"]["invalid_rows"] == 0
    assert payload["batch"]["auto_created_employees"] == [
        {
            "employee_id": 5005789865,
            "nombre": "REJON CAPILLA CARMEN DOLORES",
            "departamento": "CME-Soluciones Castro",
            "email": "emp-5005789865@pendiente.local",
            "lookup_reason": "name_search_no_match",
            "lookup_label": "Buscado por nombre sin éxito",
        }
    ]

    session = TestingSessionLocal()
    bad_departments = session.query(Department).filter(Department.name.in_([
        "01-05-2026",
        "02-05-2026",
        "03-05-2026",
        "04-05-2026",
        "05-05-2026",
        "06-05-2026",
    ])).all()
    session.close()

    assert bad_departments == []


def test_superadmin_reuses_existing_employee_when_name_and_department_match_even_if_excel_id_differs():
    seed_attendance_import_data(is_superadmin=True)
    session = TestingSessionLocal()
    existing_employee = Employee(
        id=6857890581,
        nombre="AKE GARCIA JUAN CARLOS",
        departamento="MTJ-Imagen",
        campus="Montejo",
        email="ake@example.com",
    )
    session.add(existing_employee)
    session.commit()
    session.close()

    client = TestClient(app)
    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_for_existing_employee_name_match(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["imported_rows"] == 1
    assert payload["batch"]["auto_created_employees"] == []

    session = TestingSessionLocal()
    imported_event = session.query(AttendanceEvent).filter(AttendanceEvent.nombre == "AKE GARCIA JUAN CARLOS").one()
    session.close()

    assert imported_event.employee_id == 6857890581


def test_superadmin_reuses_existing_employee_when_file_id_differs():
    seed_attendance_import_data(is_superadmin=True)
    session = TestingSessionLocal()
    session.add(
        Employee(
            id=6857890581,
            nombre="AKE GARCIA JUAN CARLOS",
            departamento="MTJ-Imagen",
            campus="Montejo",
            email="ake.garcia@example.com",
        )
    )
    session.commit()
    session.close()

    client = TestClient(app)
    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_with_existing_employee_under_different_file_id(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["auto_created_employees"] == []

    session = TestingSessionLocal()
    imported_event = session.query(AttendanceEvent).filter(AttendanceEvent.employee_id == 6857890581).one()
    reused_employee = session.query(Employee).filter(Employee.id == 3095).first()
    session.close()

    assert imported_event.nombre == "AKE GARCIA JUAN CARLOS"


def test_superadmin_reuses_existing_employee_when_name_matches_long_department_format():
    seed_attendance_import_data(is_superadmin=True)
    session = TestingSessionLocal()
    session.add(
        Employee(
            id=6857890581,
            nombre="AKE GARCIA JUAN CARLOS",
            departamento="MTJ-Imagen",
            campus="Montejo",
            email="ake.long@example.com",
        )
    )
    session.commit()
    session.close()

    client = TestClient(app)
    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_for_existing_employee_name_match_with_long_department(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["batch"]["auto_created_employees"] == []

    session = TestingSessionLocal()
    imported_events = session.query(AttendanceEvent).filter(AttendanceEvent.employee_id == 6857890581).all()
    reused_employee = session.query(Employee).filter(Employee.id == 6857890581).one()
    session.close()

    assert imported_events
    assert all(event.departamento_raw == "MTJ-Imagen" for event in imported_events)
    assert reused_employee.departamento == "MTJ-Imagen"


def test_superadmin_reuses_canonical_department_when_long_alias_belongs_to_other_department():
    seed_attendance_import_data(is_superadmin=True)
    session = TestingSessionLocal()
    conflicting_department = Department(id=77, code="cch-imagen", name="CCH-Imagen", campus="Chetumal", active=True)
    session.add(conflicting_department)
    session.flush()
    session.add(DepartmentAlias(id=78, department_id=77, alias="CCH-Imagen", source="employee"))
    session.add(DepartmentAlias(id=79, department_id=3, alias="Escuela Modelo/Chetumal/CCH-Imagen", source="employee"))
    session.commit()
    session.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "nombre",
        "ID",
        "departamento",
        "Fecha",
        "Tiempo",
        "Fuente de datos",
        "Nombre del dispositivo",
        "N.º de serie del dispositivo",
    ])
    sheet.append([
        "PERSONA CCH",
        8801,
        "Escuela Modelo/Chetumal/CCH-Imagen",
        "19-04-2026",
        "07:30",
        "Dispositivo",
        "CVA EDIFICIO PRINCIPAL",
        "FQ8321320",
    ])
    buffer = BytesIO()
    workbook.save(buffer)

    client = TestClient(app)
    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201

    session = TestingSessionLocal()
    employee = session.query(Employee).filter(Employee.id == 8801).one()
    department_links = [link.department_id for link in employee.department_links]
    session.close()

    assert employee.departamento == "CCH-Imagen"
    assert department_links == [77]


def test_superadmin_canonicalizes_long_department_names_on_auto_create():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_without_numeric_employee_ids(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201

    session = TestingSessionLocal()
    employee = session.query(Employee).filter(Employee.nombre == "NUEVA PERSONA UNO").one()
    department = session.query(Department).filter(Department.id == employee.department_links[0].department_id).one()
    imported_events = session.query(AttendanceEvent).filter(AttendanceEvent.employee_id == employee.id).all()
    aliases = {
        alias.alias
        for alias in session.query(DepartmentAlias).filter(DepartmentAlias.department_id == department.id).all()
    }
    session.close()

    assert employee.departamento == "MTJ-Imagen"
    assert employee.campus == "Montejo"
    assert department.name == "MTJ-Imagen"
    assert all(event.departamento_raw == "MTJ-Imagen" for event in imported_events)
    assert "Escuela Modelo/Montejo/MTJ-Imagen" in aliases
    assert "MTJ-Imagen" in aliases


def test_superadmin_canonicalizes_generic_escuela_modelo_prefix():
    seed_attendance_import_data(is_superadmin=True)
    client = TestClient(app)

    response = client.post(
        "/staff/attendance-imports",
        headers=auth_headers(client),
        files={
            "file": (
                "asistencia.xlsx",
                build_workbook_bytes_for_generic_prefixed_department(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201

    session = TestingSessionLocal()
    employee = session.query(Employee).filter(Employee.id == 8001).one()
    department = session.query(Department).filter(Department.id == employee.department_links[0].department_id).one()
    aliases = {
        alias.alias
        for alias in session.query(DepartmentAlias).filter(DepartmentAlias.department_id == department.id).all()
    }
    session.close()

    assert employee.departamento == "BAJAS"
    assert department.name == "BAJAS"
    assert "Escuela Modelo/BAJAS" in aliases
    assert "BAJAS" in aliases